import sys
import json
import time
import hashlib
import requests
import openpyxl
from pathlib import Path
from datetime import datetime


# 获取当前文件所在目录或完整路径, 返回 / 分隔的路径
def get_path(path_type="file"):
    if getattr(sys, "frozen", False):  # 打包后 exe 环境
        current = Path(sys.executable)
    else:  # 脚本运行环境
        current = Path(__file__).resolve()
    if path_type == "file":
        return current.as_posix()
    elif path_type == "directory":
        return current.parent.as_posix()
    else:
        raise ValueError("path_type 必须是 'file' 或 'directory'")


# 获取时间戳
def get_timestamp(format_type):
    now = datetime.now()
    if format_type == 0:
        return int(now.timestamp() * 1000)
    elif format_type == 1:
        return now.strftime("%Y/%m/%d")
    elif format_type == 2:
        return now.strftime("%Y%m%d%H%M%S")
    else:
        raise ValueError("Invalid format_type. Use 0 or 1.")


# 获取股票配置json
def get_stocks_json():
    url = "https://1024nettech.github.io/stocks/stocks.json"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        return data
    else:
        print(f"请求失败, 状态码: {response.status_code}")


# 获取股票实时点数
def get_point(url, name_value="", stock_dict={}):
    if "d.10jqka.com.cn" in url:  # 同花顺
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            response_text = response.text.strip()
            start_index = response_text.find("{")
            end_index = response_text.rfind("}") + 1
            json_str = response_text[start_index:end_index]
            data = json.loads(json_str)
            code_value = data["items"].get("5", None)
            name_value = data["items"].get("name", None)
            point_value = data["items"].get("10", None)
            return [code_value, name_value, point_value]
        else:
            print(f"同花顺请求失败, 状态码: {response.status_code}")
    elif "qt.gtimg.cn" in url:  # 腾讯财经
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            code_value = response.text.split("~")[2]
            name_value = response.text.split("~")[1]
            point_value = response.text.split("~")[3]
            return [code_value, name_value, point_value]
        else:
            print(f"腾讯财经请求失败, 状态码: {response.status_code}")
    elif "hq.sinajs.cn" in url:  # 新浪财经
        headers["Referer"] = "https://finance.sina.com.cn/"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            code_value = response.text.split("=")[0].replace("var hq_str_", "")
            name_value = response.text.split(",")[0].split('"')[1]
            point_value = response.text.split(",")[1]
            return [code_value, name_value, point_value]
        else:
            print(f"新浪财经请求失败, 状态码: {response.status_code}")
    elif "stock.xueqiu.com" in url:  # 雪球
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            code_value = data["data"][0]["symbol"]
            point_value = data["data"][0]["current"]
            url = url.replace("realtime/quotec", "quote")
            response = requests.get(url, headers=headers)
            if response.status_code == 200:
                data = response.json()
                name_value = data["data"]["quote"]["name"]
            return [code_value, name_value, point_value]
        else:
            print(f"雪球请求失败, 状态码: {response.status_code}")


# 韭圈儿解析加密字段
def split_md5(md5_str, act_time, gu_code):
    return {
        "gu_code": gu_code,
        "pe_category": "pe",
        "year": -1,
        "category": "",
        "ver": "new",
        "type": "pc",
        "version": "2.2.7",
        "authtoken": "",
        "act_time": act_time,
        "yi854tew": md5_str[29:31],
        "u54rg5d": md5_str[2:4],
        "bioduytlw": md5_str[5:6],
        "nkjhrew": md5_str[26:27],
        "bvytikwqjk": md5_str[6:8],
        "tiklsktr4": md5_str[1:2],
        "tirgkjfs": md5_str[0:2],
        "bgd7h8tyu54": md5_str[6:8],
        "yt447e13f": md5_str[8:9],
        "nd354uy4752": md5_str[30:31],
        "ghtoiutkmlg": md5_str[11:14],
        "y654b5fs3tr": md5_str[11:12],
        "fjlkatj": md5_str[2:5],
        "jnhf8u5231": md5_str[9:11],
        "sbnoywr": md5_str[23:25],
        "kf54ge7": md5_str[31:32],
        "hy5641d321t": md5_str[25:27],
        "bgiuytkw": md5_str[9:11],
        "quikgdky": md5_str[27:29],
        "ngd4uy551": md5_str[17:19],
        "bd4uy742": md5_str[26:27],
        "ngd4yut78": md5_str[12:14],
        "iogojti": md5_str[25:26],
        "h67456y": md5_str[16:19],
        "lksytkjh": md5_str[17:21],
        "n3bf4uj7y7": md5_str[18:19],
        "nbf4uj7y432": md5_str[21:23],
        "ibvytiqjek": md5_str[14:16],
        "h13ey474": md5_str[29:32],
        "abiokytke": md5_str[21:23],
        "bd24y6421f": md5_str[24:26],
        "tbvdiuytk": md5_str[16:17],
    }


# 获取股票估值分析
def get_val(url, timestamp=0, stock_dict={}, type="val"):
    if "api.jiucaishuo.com" in url:
        try:
            gu_code = stock_dict["code"]
            t = f"{timestamp}{gu_code}pepcnew2.2.7-1EWf45rlv#kfsr@k#gfksgkr"
            md5_value = hashlib.md5(t.encode("utf-8")).hexdigest()
            md5_parts = split_md5(md5_value, timestamp, gu_code)
            body = json.dumps(md5_parts)
            headers["Content-Type"] = "application/json;charset=UTF-8"
            response = requests.post(
                "https://api.jiucaishuo.com/v2/guzhi/newtubiaodata",
                headers=headers,
                data=body,
            )
            if response.status_code == 200:
                data = response.json()
                gu_name_str = data.get("data", {}).get("gu_name", "")
                point_str = (
                    data.get("data", {})
                    .get("top_data", [None, {}, {}, {}])[0]
                    .get("new_value", {})
                    .get("value", "0")
                )
                pe_str = (
                    data.get("data", {})
                    .get("top_data", [None, {}, {}, {}])[1]
                    .get("new_percent_value", {})
                    .get("value", "0")
                )
                pb_str = (
                    data.get("data", {})
                    .get("top_data", [None, {}, {}, {}])[2]
                    .get("new_percent_value", {})
                    .get("value", "0")
                )
                xilv_str = (
                    data.get("data", {})
                    .get("top_data", [None, {}, {}, {}])[3]
                    .get("new_percent_value", {})
                    .get("value", "0")
                )

                def parse_percent(s):
                    return float(s.replace("%", "").strip())

                point = parse_percent(point_str)
                pe = parse_percent(pe_str)
                pb = parse_percent(pb_str)
                xilv = parse_percent(xilv_str)
                result = (
                    pe * stock_dict["calc"][0]
                    + pb * stock_dict["calc"][1]
                    + xilv * stock_dict["calc"][2]
                )
                if type == "point":
                    return [gu_code, gu_name_str, point]
                elif type == "val":
                    return result
        except Exception as e:
            print(f"估值接口出错: {e}")
        return 0.0


# 检测并返回Excel中有数据的最后一列列号
def detect_last_col1(ws):
    for col in range(ws.max_column, 0, -1):  # 从右往左遍历
        for row in range(1, ws.max_row + 1):
            if (
                ws.cell(row=row, column=col).value is not None
                and ws.cell(row=row, column=col).value != ""
            ):
                return col
    return 1


# 数据写入Excel最后一列
def write_xlsx(results=[], xlsx_path=""):
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"文件 {xlsx_path} 未找到")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    last_col = detect_last_col1(ws)
    new_col = last_col + 1
    for i, value in enumerate(results, start=1):
        try:
            if value is None:
                ws.cell(row=i, column=new_col, value=None)
            else:
                ws.cell(row=i, column=new_col, value=round(float(value), 2))
        except ValueError:
            ws.cell(row=i, column=new_col, value=str(value))
    set_column_style(ws, new_col)
    wb.save(xlsx_path)


# 设置Excel最后一列样式
def set_column_style(ws, col_idx):
    for row in ws.iter_rows(
        min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row
    ):
        for cell in row:
            cell.font = openpyxl.styles.Font(name="宋体", size=12)
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="center", vertical="center"
            )
    ws.cell(row=2, column=col_idx).font = openpyxl.styles.Font(
        name="宋体", size=12, bold=True
    )
    for i in range(1, col_idx + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15


# ——————————————————————————————————————————————————主体代码开始——————————————————————————————————————————————————
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:148.0) Gecko/20100101 Firefox/148.0",
}
stocks = get_stocks_json()
row = 8
for k, v in stocks.items():
    if not v["point"].get("row"):
        v["point"]["row"] = row
        v["val"]["row"] = row - 1
        row += 3
results = []
for i in range(0, row - 3):
    results.append(None)
for k, v in stocks.items():
    if "api.jiucaishuo.com" in v["point"]["url"]:
        point = get_val(v["point"]["url"], get_timestamp(0), v["point"], type="point")
    else:
        point = get_point(v["point"]["url"], k, v["point"])
    results[v["point"]["row"] - 1] = point[2]
    val = get_val(v["val"]["url"], get_timestamp(0), v["val"])
    results[v["val"]["row"] - 1] = val
    time.sleep(1)
results[0] = get_timestamp(1)
results[1] = "上证"
write_xlsx(results, Path(get_path("directory")) / "stocks_data.xlsx")
# End-287-2025.12.15.094645

